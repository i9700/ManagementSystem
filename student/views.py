from django.shortcuts import render, HttpResponse, redirect
from .models import *
from django.db.models import F, Q
from openpyxl import load_workbook
import os


# Create your views here.

def index(request):
    student_list = Student.objects.all()
    class_list = Clas.objects.all()
    return render(request, "student/index.html", {"student_list": student_list, "class_list": class_list})


def add_student(request):
    if request.method == "GET":
        class_list = Clas.objects.all()
        course_list = Course.objects.all()

        return render(request, "student/add_student.html", {"class_list": class_list, "course_list": course_list})
    else:
        print(request.POST)
        print(request.POST.dict())

        Student.objects.create(**request.POST.dict())
        return redirect("/student/index")


def delete_student(request, del_id):
    student = Student.objects.get(pk=del_id)
    if request.method == "GET":
        return render(request, "student/del_student.html", {"student": student})
    else:
        student.delete()
        return redirect("/student/index")


def edit_student(request, edit_id):
    student = Student.objects.get(pk=edit_id)
    class_list = Clas.objects.all()
    course_list = Course.objects.all()
    if request.method == "GET":
        return render(request, "student/edit_student.html",
                      {"student": student, "class_list": class_list, "course_list": course_list})
    else:
        course_id_list = request.POST.getlist("course_id_list")  # ['1', '5']
        data = request.POST.dict()
        data.pop("course_id_list")
        Student.objects.filter(pk=edit_id).update(**data)
        student.course.set(course_id_list)
        return redirect("/student/index")


def elective(request):
    course_list = Course.objects.all()

    print(request.POST)
    if request.method == "GET":
        return render(request, "student/course.html", {"course_list": course_list})
    else:
        student_id = 3
        student = Student.objects.get(pk=student_id)
        course_id_list = request.POST.getlist("course_id_list")
        student.course.set(course_id_list)
        return redirect("/student/index")


def search(request):
    class_list = Clas.objects.all()
    print(request.POST)
    key_word = request.POST.dict().get("key_word")
    class_id = request.POST.dict().get("class_id")
    print(key_word)
    print(class_id, type(class_id))
    if class_id == "" and key_word == "":
        return redirect("/student/index")
    if class_id and key_word:
        student_list = Student.objects.filter(Q(name__contains=key_word) & Q(clas_id=class_id))
        return render(request, "student/index.html",
                      {"student_list": student_list, "class_list": class_list, "key_word": key_word,
                       "class_id": int(class_id)})
    else:
        if not key_word:
            student_list = Student.objects.filter(clas_id=class_id)
            return render(request, "student/index.html",
                          {"student_list": student_list, "class_list": class_list, "key_word": key_word,
                           "class_id": int(class_id)})
        else:
            student_list = Student.objects.filter(name__contains=key_word)
            return render(request, "student/index.html",
                          {"student_list": student_list, "class_list": class_list, "key_word": key_word})


def stu_excel(request):
    file_obj = request.FILES.get("stu_excel")

    # (1)将上传文件下载到服务器某个文件夹下:

    path = os.path.join('media', 'files', file_obj.name)

    with open(path, "wb") as f:
        for line in file_obj:
            f.write(line)

    # (2) 通过python操作excel表格
    wb = load_workbook(path)
    work_sheet = wb.worksheets[0]  # 获取第一个sheet对象

    student_list = []
    for line in work_sheet.iter_rows(min_row=3):
        print(line)
        if not line[0].value:
            break

        # 学生详情记录
        stu_detail = StudentDetail.objects.create(tel=line[4].value, addr=line[5].value)
        class_id = Clas.objects.get(name=line[6].value).id
        sex = line[2].value
        if sex == "男":
            sex = 0
        else:
            sex = 1
        stu = Student(name=line[0].value,
                      age=line[1].value,
                      sex=sex,
                      birthday=line[3].value,
                      stu_detail=stu_detail,
                      clas_id=class_id,
                      )
        student_list.append(stu)
    Student.objects.bulk_create(student_list)  # 批量创建，减少撞库
    return redirect("/student/index")
